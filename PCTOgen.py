import argparse
from docxtpl import DocxTemplate
from openpyxl import load_workbook
import subprocess
import log
from pyfiglet import Figlet
import os.path

VERSION = 1.0
DOCX_TEMPLATES = [
    '02-Convenzione_studente_per_stage.docx',
    '03-Patto_formativo_studente.docx',
    '04-Progetto_formativo_studente.docx',
    '05-Foglio_presenze.docx',
    '06-Scheda_di_valutazione_studente.docx',
    '07-Modulo_di_gradimento_studente.docx',
    '08-Relazione_di_fine_attività.docx'
]

def student_to_path(student):
    return student.replace(" ", "_")

def render_pdf(in_path, out_path):
    subprocess.call(['soffice',
                 '--headless',
                 '--convert-to',
                 'pdf',
                 '--outdir',
                 out_path,
                 in_path])


def render_PCTO_pdf(studente):
    log.info("Generating final PCTO document file for student {}".format(studente)) 
    os.system("pdftk " + studente + "/*.pdf cat output " + studente + "/" + studente + "_PCTO_TO_PRINT.PDF")

def render_docx(studente, fpath, context):
    if not os.path.exists(fpath):
        log.error("File {} non esiste".format(fpath))
        quit()  
    doc = DocxTemplate(fpath)
    doc.render(context)
    doc_dir = student_to_path(studente)
    if not os.path.exists(doc_dir):
        os.mkdir(doc_dir)
    doc_path = doc_dir + "/" + student_to_path(studente) + "_" + fpath
    doc.save(doc_path)
    log.info("Generating {} file for student {}".format(fpath, studente))  
    render_pdf(doc_path, studente.replace(" ", "_"))
    log.info("Rendering PDF version of {} document for student {}".format(fpath, studente))  

def read_validate_PCTODB(fpath):
    if not os.path.exists(fpath):
        log.error("File {} non esiste".format(fpath))
        quit()
    
    wb = load_workbook(fpath, read_only=True)
    ws = wb.sheetnames
    
    if "PCTO" not in ws:
        log.error("Il worksheet PCTO non è presente")
        quit()
    if "AZIENDA" not in ws:
        log.error("Il worksheet AZIENDA non è presente")
        quit()
    if "STUDENTI" not in ws:
        log.error("Il worksheet STUDENTI non è presente")
        quit()
    log.info("File {} caricato correttamente".format(fpath))
    return wb
    
def get_ws_cells(ws, col, rows, rowe):
    # col is the column of the key cells
    # rows is the start row of key and value cells
    # rowe is the end row  of key and value cells
    # col + 1 is the column of the value cells

    cells = {}
    for row in range(rows, rowe + 1):
        if ws.cell(column=col+1, row=row).value is not None:
            cells[ws.cell(column=col, row=row).value] = ws.cell(column=col+1, row=row).value
    return cells

def main(args):
    log.init() 

    # print banner
    f = Figlet(font='small')
    print(f.renderText('PCTOgen v {}'.format(VERSION)))
    log.info("PCTOgen running!!!")
    
    # read and validate PCTO DB file
    wb = read_validate_PCTODB(args.i)

    # extract cells from DB
    log.info("Extracting cells from DB file")
    pcto = get_ws_cells(wb['PCTO'], col = 1, rows = 3, rowe = 21)
    azienda = get_ws_cells(wb['AZIENDA'], col = 1, rows = 3, rowe = 12)
    studenti_misc = get_ws_cells(wb['STUDENTI'], col = 1, rows = 3, rowe = 4)
    studenti_names = get_ws_cells(wb['STUDENTI'], col = 1, rows = 8, rowe = 27)
    context = pcto | azienda | studenti_misc
    #Generates documents package for each student
    for student in studenti_names:
        studente = studenti_names[student] 
        log.info("")
        log.info("Generating PCTO documentation for {}".format(studente))
        context['STUDENTE'] = studente 
        for doc in DOCX_TEMPLATES:
            render_docx(studente, doc, context)
        render_PCTO_pdf(student_to_path(studente))



if __name__ == "__main__":
    parser = argparse.ArgumentParser(
                      description='Automatically generate PCTO documentation \
                                   package starting from an excel file. Libreoffice and pdftk are needed')
    parser.add_argument('-i',
                        help='Input PCTO DB file in .xlsx format', required=True)
   
    args = parser.parse_args()
    main(args)



